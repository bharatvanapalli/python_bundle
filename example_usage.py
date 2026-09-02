"""
Example script demonstrating openpyxl usage.
Supports both:
1. Normal Python environment import (if installed via pip / wheels)
2. Vendored directory fallback (sys.path) if pip install cannot be run
"""
import sys
from pathlib import Path

# Fallback: if openpyxl is not installed globally/in venv, load from vendor directory
try:
    import openpyxl
except ImportError:
    vendor_dir = Path(__file__).resolve().parent / "vendor"
    if vendor_dir.exists():
        sys.path.insert(0, str(vendor_dir))
    import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def main():
    print(f"Using openpyxl version: {openpyxl.__version__}")
    
    # 1. Create a new workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Offline Bundle Demo"

    # 2. Add header formatting
    headers = ["ID", "Name", "Category", "Quantity", "Unit Price", "Total Price"]
    ws.append(headers)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 3. Add sample data
    rows = [
        [101, "Laptop", "Electronics", 5, 750.00, "=D2*E2"],
        [102, "Desk Chair", "Furniture", 12, 120.00, "=D3*E3"],
        [103, "Wireless Mouse", "Electronics", 25, 25.50, "=D4*E4"],
        [104, "Standing Desk", "Furniture", 4, 350.00, "=D5*E5"],
    ]

    for row in rows:
        ws.append(row)

    # 4. Format numbers and borders
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border
            if cell.column in [4, 5, 6]:
                cell.alignment = Alignment(horizontal="right")
            if cell.column in [5, 6]:
                cell.number_format = '$#,##0.00'

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 5. Save workbook
    output_filename = "sample_output.xlsx"
    wb.save(output_filename)
    print(f"Successfully generated Excel workbook: {output_filename}")

    # 6. Read back workbook to verify
    wb_read = openpyxl.load_workbook(output_filename, data_only=False)
    sheet_read = wb_read.active
    print(f"Read back sheet name: '{sheet_read.title}', Total Rows: {sheet_read.max_row}, Total Columns: {sheet_read.max_column}")
    print("Verification complete! openpyxl is fully functional.")

if __name__ == "__main__":
    main()
